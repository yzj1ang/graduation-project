import os
from numpy.lib.function_base import average
import openpyxl
from openpyxl import workbook
import numpy as np

ROOT_FILE = "./reports"
RAND1 = "./reports/randSeed=1"
XLSX = "./result.xlsx"
SHEETNAME_MAP = {}
SHEETNAME_MAP['Epidemic'] = "EpidemicRouter"
SHEETNAME_MAP['Degree'] = "DegreeCentralityRouter"
SHEETNAME_MAP['Close'] = "ClosenessCentralityRouter"
SHEETNAME_MAP['Between'] = "BetweennessCentralityRouter"
SHEETNAME_MAP['BRidging'] = "BRidgingCentralityRouter"

RANDSEED_MAP = {}
RANDSEED_MAP['1'] = 0
RANDSEED_MAP['3'] = 1
RANDSEED_MAP['7'] = 2
RANDSEED_MAP['11'] = 3
RANDSEED_MAP['13'] = 4
RANDSEED_MAP['17'] = 5
RANDSEED_MAP['19'] = 6

TTL_MAP = {}
TTL_MAP['60'] = 0
TTL_MAP['120'] = 1
TTL_MAP['180'] = 2
TTL_MAP['240'] = 3
TTL_MAP['300'] = 4
TTL_MAP['360'] = 5

ROW_MAP = {}
ROW_MAP['sim_time'] = 0
ROW_MAP['created'] = 1
ROW_MAP['started'] = 2
ROW_MAP['relayed'] = 3
ROW_MAP['aborted'] = 4
ROW_MAP['dropped'] = 5
ROW_MAP['removed'] = 6
ROW_MAP['delivered'] = 7
ROW_MAP['delivery_prob'] = 8
ROW_MAP['response_prob'] = 9
ROW_MAP['overhead_ratio'] = 10
ROW_MAP['latency_avg'] = 11
ROW_MAP['latency_med'] = 12
ROW_MAP['hopcount_avg'] = 13
ROW_MAP['hopcount_med'] = 14
ROW_MAP['buffertime_avg'] = 15
ROW_MAP['buffertime_med'] = 16
ROW_MAP['rtt_avg'] = 17
ROW_MAP['rtt_med'] = 18

def process_file():
    for root, dirs, files in os.walk(RAND1):
        print(root)  # 当前目录路径（包含所有子目录）
        # print("===============")
        # print(dirs) #当前路径下所有子目录（同一路径下的存一个列表中）
        # print("===============")
        # print(files) #当前路径下所有非目录子文件（同一路径下的存一个列表中）
        for filename in files:
            filedir = RAND1 + "/" + filename
            print("processing: "+filedir+"...", end='')
            with open(filedir, "r+") as fp:
                fp.readline()
                lines = fp.readlines()
                result = []
                append = True
                for linecnotent in lines:
                    lineitem = linecnotent.split(": ")
                    if len(lineitem)==2 :
                        result.append(lineitem[1])
                    else:
                        append = False

                if append :
                    for i in result:
                        fp.write(i)
            print("Done!", end='\n')

def genreate_xlsx():
    # workbook = openpyxl.Workbook()
    workbook = openpyxl.load_workbook(XLSX)

    for j in ['n27', 'n100']:
        for i in SHEETNAME_MAP.keys():
            if j+'_'+i in workbook.sheetnames:
                continue

            workbook.create_sheet(j+'_'+i)
    workbook.create_sheet('n27')
    workbook.create_sheet('n100')
    workbook.save(XLSX)
    workbook.close()

def clean_xlsx():
    workbook = openpyxl.load_workbook(XLSX)

    for j in ['n27', 'n100']:
        for i in SHEETNAME_MAP.keys():
            del workbook[j+'_'+i]

    workbook.close()
    del workbook['n27']
    del workbook['n100']
    workbook.save(XLSX)


def get_data(type='n27', router="EpidemicRouter"):
    target_router = [[[] for j in range(6)] for k in range(7)]

    for root, dirs, files, in os.walk(ROOT_FILE):
        # only process the simulation result from the root directory
        if root != ROOT_FILE:
            continue

        # iterate all the files in root directory
        for filename in files:
            # if simulation result is valid
            name_list = filename.split("_")
            if len(name_list) != 6:
                continue

            # decide the n type: `n27` or `n100`
            if name_list[1] != type:
                continue

            # decide the type of router: "EpidemicRouter", "DegreeCentralityRouter", "ClosenessCentralityRouter",
            # "BetweennessCentralityRouter", "BRidgingCentralityRouter"
            if name_list[2] != router:
                continue

            # get the simulation info : which `ttl` and `randSeed`
            ttl = TTL_MAP[name_list[3]]
            rand_seed = RANDSEED_MAP[name_list[4]]

            filedir = ROOT_FILE + "/" + filename
            with open(filedir, "r+") as fp:
                fp.readline()
                lines = fp.readlines()
                for linecnotent in lines:
                    linecnotent = linecnotent.replace("\n", '')
                    lineitem = linecnotent.split(": ")
                    if len(lineitem) == 2:
                        content = lineitem[1]
                        if lineitem[1].isdigit:
                            content = float(lineitem[1])
                        target_router[rand_seed][ttl].append(content)

    return target_router

def get_average(type='n27', router="EpidemicRouter"):
    # get the specific data
    target_data = get_data(type, router)

    # remove the nonnumeric data
    for rand_set in target_data:
        for ttl_set in rand_set:
            del ttl_set[-1]
            del ttl_set[-1]

    # use numpy to get the average data
    np_data = np.array(target_data, dtype=np.float64)
    ave_data = np.average(np_data, axis=0).tolist()
    return ave_data

def write_each_xlsx():
    workbook = openpyxl.load_workbook(XLSX)

    for i in workbook.sheetnames:
        sheet_name = i.split('_')
        if len(sheet_name)!=2:
            continue
        sheet = workbook[i]
        sheet_row = 1

        # write the head of the table
        for index, head in enumerate(TTL_MAP.keys(), start=2):
            sheet.cell(row=sheet_row, column=index, value=int(head))

        # get the simulation data from specific n type and router
        target_router = get_data(type=sheet_name[0], router=SHEETNAME_MAP[sheet_name[1]])

        for index, rand_set in enumerate(target_router):
            # reset the column to 1 and write the rand seed of the data section
            sheet_col = 1
            sheet.cell(row=sheet_row, column=sheet_col, value="Rand Seed = "+list(RANDSEED_MAP.keys())[index])
            sheet_row += 1

            # write the row label of the table
            for index, row_label in enumerate(ROW_MAP.keys(), start=sheet_row):
                sheet.cell(row=index, column=sheet_col, value=row_label)

            # write the each data item from the ttl set
            for ttl_set in rand_set:
                sheet_col += 1
                for index, item in enumerate(ttl_set, start=sheet_row):
                    sheet.cell(row=index, column=sheet_col, value=item)

            # switch a step
            sheet_row +=20

    workbook.save(XLSX)
    workbook.close()

def write_ave_xlsx():
    workbook = openpyxl.load_workbook(XLSX)

    for n_type in ['n27', 'n100']:
        sheet = workbook[n_type]
        sheet_row = 1

        for index, head in enumerate(TTL_MAP.keys(), start=2):
            sheet.cell(row=sheet_row, column=index, value=head)

        for router_type in SHEETNAME_MAP.values():
            average_data = get_average(type=n_type, router=router_type)

            # reset the column to 1 and write the router type of the data section
            sheet_col = 1
            sheet.cell(row=sheet_row, column=sheet_col, value=router_type)
            sheet_row += 1

            # write the row label of the table
            for index, row_label in enumerate(ROW_MAP.keys(), start=sheet_row):
                if row_label == 'rtt_avg':
                    break
                sheet.cell(row=index, column=sheet_col, value=row_label)

            # write the each data item from the ttl set
            for ttl_set in average_data:
                sheet_col += 1
                for index, item in enumerate(ttl_set, start=sheet_row):
                    sheet.cell(row=index, column=sheet_col, value=item)

            # switch a step
            sheet_row += 18

    workbook.save(XLSX)
    workbook.close()

if __name__=="__main__":
    # clean_xlsx()
    genreate_xlsx()
    write_each_xlsx()
    write_ave_xlsx()
